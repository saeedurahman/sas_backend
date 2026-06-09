"""CSV and Excel export services."""

import csv
import io
from datetime import date
from decimal import Decimal
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

_SALES_HEADERS = [
    "Sale Number",
    "Date",
    "Branch",
    "Customer",
    "Sale Type",
    "Status",
    "Subtotal",
    "Discount",
    "Tax",
    "Total Amount",
    "Total Paid",
    "Balance Due",
    "Payment Methods",
    "Cashier",
]

_INVENTORY_HEADERS = [
    "Product",
    "Variation",
    "Category",
    "Branch",
    "Current Qty",
    "Unit",
    "Cost Price",
    "Total Value",
]


def _dec(value) -> Decimal:
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


def _branch_filter(branch_id: UUID | None) -> str:
    if branch_id is None:
        return ""
    return " AND s.branch_id = :branch_id"


def _inventory_branch_filter(branch_id: UUID | None) -> str:
    if branch_id is None:
        return ""
    return " AND sq.branch_id = :branch_id"


async def _fetch_sales_rows(
    db: AsyncSession,
    business_id: UUID,
    date_from: date,
    date_to: date,
    branch_id: UUID | None = None,
) -> list[dict]:
    branch_sql = _branch_filter(branch_id)
    params: dict = {
        "business_id": business_id,
        "date_from": date_from,
        "date_to": date_to,
    }
    if branch_id is not None:
        params["branch_id"] = branch_id

    sql = f"""
        SELECT
            s.sale_number,
            s.sold_at,
            b.name AS branch_name,
            COALESCE(c.name, '') AS customer_name,
            s.sale_type,
            s.status,
            COALESCE(SUM(sl.qty * sl.unit_price), 0) AS subtotal,
            COALESCE(SUM(sl.discount_amount), 0) AS discount,
            COALESCE(SUM(sl.tax_amount), 0) AS tax,
            COALESCE(SUM(sl.qty * sl.unit_price - sl.discount_amount + sl.tax_amount), 0)
                AS total_amount,
            COALESCE((
                SELECT SUM(sp.amount)
                FROM sale_payments sp
                WHERE sp.sale_id = s.id
                    AND sp.business_id = s.business_id
                    AND sp.deleted_at IS NULL
                    AND sp.status = 'completed'
            ), 0) AS total_paid,
            COALESCE((
                SELECT STRING_AGG(DISTINCT sp.payment_method::text, ', ')
                FROM sale_payments sp
                WHERE sp.sale_id = s.id
                    AND sp.business_id = s.business_id
                    AND sp.deleted_at IS NULL
                    AND sp.status = 'completed'
            ), '') AS payment_methods,
            u.full_name AS cashier_name
        FROM sales s
        INNER JOIN branches b ON b.id = s.branch_id
        LEFT JOIN customers c ON c.id = s.customer_id
        INNER JOIN sale_lines sl ON sl.sale_id = s.id
            AND sl.business_id = s.business_id
            AND sl.deleted_at IS NULL
        INNER JOIN users u ON u.id = s.created_by
        WHERE s.business_id = :business_id
            AND s.deleted_at IS NULL
            AND s.sold_at::date BETWEEN :date_from AND :date_to
            {branch_sql}
        GROUP BY
            s.id, s.sale_number, s.sold_at, b.name, c.name,
            s.sale_type, s.status, u.full_name
        ORDER BY s.sold_at DESC
    """
    result = await db.execute(text(sql), params)
    rows: list[dict] = []
    for row in result:
        total_amount = _dec(row.total_amount)
        total_paid = _dec(row.total_paid)
        balance_due = max(total_amount - total_paid, Decimal("0"))
        rows.append(
            {
                "sale_number": row.sale_number,
                "sold_at": row.sold_at,
                "branch_name": row.branch_name,
                "customer_name": row.customer_name,
                "sale_type": row.sale_type,
                "status": row.status,
                "subtotal": _dec(row.subtotal),
                "discount": _dec(row.discount),
                "tax": _dec(row.tax),
                "total_amount": total_amount,
                "total_paid": total_paid,
                "balance_due": balance_due,
                "payment_methods": row.payment_methods or "",
                "cashier_name": row.cashier_name,
            }
        )
    return rows


def _sales_row_to_list(row: dict) -> list:
    sold_at = row["sold_at"]
    date_str = sold_at.strftime("%Y-%m-%d %H:%M") if sold_at else ""
    return [
        row["sale_number"],
        date_str,
        row["branch_name"],
        row["customer_name"],
        row["sale_type"],
        row["status"],
        str(row["subtotal"]),
        str(row["discount"]),
        str(row["tax"]),
        str(row["total_amount"]),
        str(row["total_paid"]),
        str(row["balance_due"]),
        row["payment_methods"],
        row["cashier_name"],
    ]


async def export_sales_csv(
    db: AsyncSession,
    business_id: UUID,
    date_from: date,
    date_to: date,
    branch_id: UUID | None = None,
) -> str:
    rows = await _fetch_sales_rows(
        db, business_id, date_from, date_to, branch_id
    )
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(_SALES_HEADERS)
    for row in rows:
        writer.writerow(_sales_row_to_list(row))
    return buffer.getvalue()


async def export_sales_excel(
    db: AsyncSession,
    business_id: UUID,
    date_from: date,
    date_to: date,
    branch_id: UUID | None = None,
) -> bytes:
    rows = await _fetch_sales_rows(
        db, business_id, date_from, date_to, branch_id
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    bold = Font(bold=True)
    for col_idx, header in enumerate(_SALES_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(_sales_row_to_list(row), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def _fetch_inventory_rows(
    db: AsyncSession,
    business_id: UUID,
    branch_id: UUID | None = None,
) -> list[dict]:
    branch_sql = _inventory_branch_filter(branch_id)
    params: dict = {"business_id": business_id}
    if branch_id is not None:
        params["branch_id"] = branch_id

    sql = f"""
        WITH stock_qty AS (
            SELECT
                sm.branch_id,
                sm.product_id,
                sm.variation_id,
                COALESCE(SUM(sm.qty), 0) AS current_qty
            FROM stock_movements sm
            WHERE sm.business_id = :business_id
                AND sm.deleted_at IS NULL
                {branch_sql.replace("sq.", "sm.")}
            GROUP BY sm.branch_id, sm.product_id, sm.variation_id
        ),
        avg_costs AS (
            SELECT
                pl.product_id,
                pl.variation_id,
                AVG(pl.cost_per_unit) AS avg_cost
            FROM purchase_lines pl
            WHERE pl.business_id = :business_id
                AND pl.deleted_at IS NULL
                AND pl.qty_remaining > 0
            GROUP BY pl.product_id, pl.variation_id
        )
        SELECT
            p.name AS product_name,
            pv.name AS variation_name,
            cat.name AS category_name,
            b.name AS branch_name,
            sq.current_qty,
            COALESCE(u.symbol, bu.symbol) AS unit_symbol,
            COALESCE(ac.avg_cost, 0) AS cost_price,
            sq.current_qty * COALESCE(ac.avg_cost, 0) AS total_value
        FROM stock_qty sq
        INNER JOIN products p ON p.id = sq.product_id
        INNER JOIN branches b ON b.id = sq.branch_id
        LEFT JOIN categories cat ON cat.id = p.category_id
        LEFT JOIN product_variations pv ON pv.id = sq.variation_id
        LEFT JOIN units u ON u.id = pv.unit_id
        LEFT JOIN units bu ON bu.id = p.base_unit_id
        LEFT JOIN avg_costs ac ON ac.product_id = sq.product_id
            AND ac.variation_id IS NOT DISTINCT FROM sq.variation_id
        ORDER BY p.name, pv.name NULLS FIRST, b.name
    """
    result = await db.execute(text(sql), params)
    return [
        {
            "product_name": row.product_name,
            "variation_name": row.variation_name or "",
            "category_name": row.category_name or "",
            "branch_name": row.branch_name,
            "current_qty": _dec(row.current_qty),
            "unit_symbol": row.unit_symbol or "",
            "cost_price": _dec(row.cost_price),
            "total_value": _dec(row.total_value),
        }
        for row in result
    ]


def _inventory_row_to_list(row: dict) -> list:
    return [
        row["product_name"],
        row["variation_name"],
        row["category_name"],
        row["branch_name"],
        str(row["current_qty"]),
        row["unit_symbol"],
        str(row["cost_price"]),
        str(row["total_value"]),
    ]


async def export_inventory_csv(
    db: AsyncSession,
    business_id: UUID,
    branch_id: UUID | None = None,
) -> str:
    rows = await _fetch_inventory_rows(db, business_id, branch_id)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(_INVENTORY_HEADERS)
    for row in rows:
        writer.writerow(_inventory_row_to_list(row))
    return buffer.getvalue()


async def export_inventory_excel(
    db: AsyncSession,
    business_id: UUID,
    branch_id: UUID | None = None,
) -> bytes:
    rows = await _fetch_inventory_rows(db, business_id, branch_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    bold = Font(bold=True)
    for col_idx, header in enumerate(_INVENTORY_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(_inventory_row_to_list(row), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
